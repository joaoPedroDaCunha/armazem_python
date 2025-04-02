import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinter import messagebox

def salvar_dados(date,horario,nome,telefone,placa,tipo,trans,forn,prod,carga,val,nf1,nfpalete1,qtd1,lote1,peso1,checkbox_lote2_var,nf2,nfpalete2,qtd2,lote2,peso2,checkbox_lote3_var,nf3,nfpalete3,qtd3,lote3,peso3):
    try:
        if date and horario and nome and telefone and placa and tipo and trans and forn and carga and prod and nf1 and lote1 and peso1:
            # Cria dicionário com os dados principais
            dados_programacao = {
                'Data de Entrada': [date], 
                'Horario de entrada': [horario], 
                'Nome do motorista': [nome],
                'Telefone': [telefone],
                'NF': [nf1],
                'Fornecedor': [forn],
                'Peso total': [peso1],
                'Placa': [placa],
                'Tipo de veiculo': [tipo],
                'Tipo do produto': [prod + " " + carga],
                'Transportadora': [trans]
            }

            # Converte o dicionário em DataFrame
            df_programacao = pd.DataFrame(dados_programacao)

            # Cria lista com dados do primeiro lote
            dados_planilha = [
                {'Movimento': "ENTRADA", 'EMISSÃO NF': date, 'Placa': placa, 'Tipo de veiculo': tipo, 'Transportadora': trans, 'Material': prod, 'Tipo de Carga': carga, 'Fornecedor': forn, 'NF fornecedor': nf1, 'NF Palete': nfpalete1, 'QT NF palete': qtd1, 'Lote do fornecedor': lote1, 'Validade': val, 'Peso': peso1}
            ]

            # Converte a lista em DataFrame
            df_planilha = pd.DataFrame(dados_planilha)

            # Verifica se o lote 2 deve ser incluído
            if checkbox_lote2_var.get() == 1:
                if nf2 and lote2 and peso2:
                    df_planilha = pd.concat([df_planilha, pd.DataFrame([{'Movimento': "ENTRADA", 'EMISSÃO NF': date, 'Placa': placa, 'Tipo de veiculo': tipo, 'Transportadora': trans, 'Material': prod, 'Tipo de Carga': carga, 'Fornecedor': forn, 'NF fornecedor': nf2, 'NF Palete': nfpalete2, 'QT NF palete': qtd2, 'Lote do fornecedor': lote2, 'Validade': val, 'Peso': peso2}])], ignore_index=True)

            # Verifica se o lote 3 deve ser incluído
            if checkbox_lote3_var.get() == 1:
                if nf3 and lote3 and peso3:
                    df_planilha = pd.concat([df_planilha, pd.DataFrame([{'Movimento': "ENTRADA", 'EMISSÃO NF': date, 'Placa': placa, 'Tipo de veiculo': tipo, 'Transportadora': trans, 'Material': prod, 'Tipo de Carga': carga, 'Fornecedor': forn, 'NF fornecedor': nf3, 'NF Palete': nfpalete3, 'QT NF palete': qtd3, 'Lote do fornecedor': lote3, 'Validade': val, 'Peso': peso3}])], ignore_index=True)

            # Verifica se o arquivo já existe
            if os.path.exists('dados.xlsx'):
                # Abre o arquivo existente e adiciona os dados
                wb = load_workbook('dados.xlsx')

                # Adiciona os dados à planilha de programação
                if 'Programacao' not in wb.sheetnames:
                    ws_programacao = wb.create_sheet("Programacao")
                else:
                    ws_programacao = wb['Programacao']

                for row in dataframe_to_rows(df_programacao, index=False, header=False):
                    ws_programacao.append(row)

                # Adiciona os dados à planilha de lote
                if 'Planilha' not in wb.sheetnames:
                    ws_planilha = wb.create_sheet("Planilha")
                else:
                    ws_planilha = wb['Planilha']

                for row in dataframe_to_rows(df_planilha, index=False, header=False):
                    ws_planilha.append(row)

                # Adiciona os dados à planilha de descarga de sal (exemplo)
                if 'Descarga do Sal' not in wb.sheetnames:
                    ws_descarga_sal = wb.create_sheet("Descarga do Sal")
                else:
                    ws_descarga_sal = wb['Descarga do Sal']
                    ws_descarga_sal['D8'] = date
                    ws_descarga_sal['K8'] = horario
                    ws_descarga_sal['D10'] = nome
                    ws_descarga_sal['D12'] = telefone
                    ws_descarga_sal['D14'] = tipo
                    ws_descarga_sal['K14'] = placa
                    ws_descarga_sal['D16'] = trans
                    ws_descarga_sal['D18'] = forn
                    ws_descarga_sal['M18'] = carga
                    ws_descarga_sal['D20'] = nf1
                    ws_descarga_sal['K20'] = nfpalete1
                    ws_descarga_sal['P20'] = peso1
                    if checkbox_lote2_var.get() == 1:
                        if nf1 == nf2 :
                            ws_descarga_sal['D22'] = " "
                            ws_descarga_sal['K22'] = " "
                            ws_descarga_sal['P22'] = " "
                            ws_descarga_sal['P20'] = peso1+peso2
                        else :
                            ws_descarga_sal['D22'] = nf2
                            ws_descarga_sal['K22'] = nfpalete2
                            ws_descarga_sal['P22'] = peso2
                    else :
                        ws_descarga_sal['D22'] = " "
                        ws_descarga_sal['K22'] = " "
                        ws_descarga_sal['P22'] = " "
                    if checkbox_lote3_var.get() == 1:
                        if nf1 == nf3 and nf1 == nf2:
                            ws_descarga_sal['D24'] = " "
                            ws_descarga_sal['K24'] = " "
                            ws_descarga_sal['P24'] = " "
                            ws_descarga_sal['P20'] = peso1+peso2+peso3
                        else:
                            ws_descarga_sal['D24'] = nf3
                            ws_descarga_sal['K24'] = nfpalete3
                            ws_descarga_sal['P24'] = peso3
                    else :
                        ws_descarga_sal['D24'] = " "
                        ws_descarga_sal['K24'] = " "
                        ws_descarga_sal['P24'] = " "
                    ws_descarga_sal['D26'] = prod
                    ws_descarga_sal['L26'] = val
                    ws_descarga_sal['D28'] = lote1
                    ws_descarga_sal['H28'] = nf1
                    ws_descarga_sal['K28'] = peso1
                    ws_descarga_sal['O28'] = qtd1
                    if checkbox_lote2_var.get() == 1:
                        ws_descarga_sal['D30'] = lote2
                        ws_descarga_sal['H30'] = nf2
                        ws_descarga_sal['K30'] = peso2
                        ws_descarga_sal['O30'] = qtd2
                    else :
                        ws_descarga_sal['D30'] = " "
                        ws_descarga_sal['H30'] = " "
                        ws_descarga_sal['K30'] = " "
                        ws_descarga_sal['O30'] = " "
                    if checkbox_lote3_var.get() == 1:
                        ws_descarga_sal['D32'] = lote3
                        ws_descarga_sal['H32'] = nf3
                        ws_descarga_sal['K32'] = peso3
                        ws_descarga_sal['O32'] = qtd3
                    else :
                        ws_descarga_sal['D32'] = " "
                        ws_descarga_sal['H32'] = " "
                        ws_descarga_sal['K32'] = " "
                        ws_descarga_sal['O32'] = " "


            else:
                # Se o arquivo não existir, cria um novo
                wb = Workbook()

                # Cria a planilha de Programacao
                ws_programacao = wb.active
                ws_programacao.title = "Programacao"
                for row in dataframe_to_rows(df_programacao, index=False, header=True):
                    ws_programacao.append(row)

                # Cria a planilha de Planilha
                ws_planilha = wb.create_sheet("Planilha")
                for row in dataframe_to_rows(df_planilha, index=False, header=True):
                    ws_planilha.append(row)

                # Cria a planilha de Descarga do Sal
                ws_descarga_sal = wb.create_sheet("Descarga do Sal")
                ws_descarga_sal = wb['Descarga do Sal']
                ws_descarga_sal['D8'] = date
                ws_descarga_sal['K8'] = horario
                ws_descarga_sal['D10'] = nome
                ws_descarga_sal['D12'] = telefone
                ws_descarga_sal['D14'] = tipo
                ws_descarga_sal['K14'] = placa
                ws_descarga_sal['D16'] = trans
                ws_descarga_sal['D18'] = forn
                ws_descarga_sal['M18'] = carga
                ws_descarga_sal['D20'] = nf1
                ws_descarga_sal['K20'] = nfpalete1
                ws_descarga_sal['P20'] = peso1
                ws_descarga_sal['D20'] = nf1
                ws_descarga_sal['K20'] = nfpalete1
                ws_descarga_sal['P20'] = peso1
                if checkbox_lote2_var.get() == 1:
                    if nf1 == nf2 :
                        ws_descarga_sal['D22'] = " "
                        ws_descarga_sal['K22'] = " "
                        ws_descarga_sal['P22'] = " "
                        ws_descarga_sal['P20'] = peso1+peso2
                    else :
                        ws_descarga_sal['D22'] = nf2
                        ws_descarga_sal['K22'] = nfpalete2
                        ws_descarga_sal['P22'] = peso2
                else :
                    ws_descarga_sal['D22'] = " "
                    ws_descarga_sal['K22'] = " "
                    ws_descarga_sal['P22'] = " "
                if checkbox_lote3_var.get() == 1:
                    if nf1 == nf3 and nf1 == nf2:
                        ws_descarga_sal['D24'] = " "
                        ws_descarga_sal['K24'] = " "
                        ws_descarga_sal['P24'] = " "
                        ws_descarga_sal['P20'] = peso1+peso2+peso3
                    else:
                        ws_descarga_sal['D24'] = nf3
                        ws_descarga_sal['K24'] = nfpalete3
                        ws_descarga_sal['P24'] = peso3
                else :
                    ws_descarga_sal['D24'] = " "
                    ws_descarga_sal['K24'] = " "
                    ws_descarga_sal['P24'] = " "
                ws_descarga_sal['D26'] = prod
                ws_descarga_sal['L26'] = val
                ws_descarga_sal['D28'] = lote1
                ws_descarga_sal['H28'] = nf1
                ws_descarga_sal['K28'] = peso1
                ws_descarga_sal['O28'] = qtd1
                if checkbox_lote2_var.get() == 1:
                    ws_descarga_sal['D30'] = lote2
                    ws_descarga_sal['H30'] = nf2
                    ws_descarga_sal['K30'] = peso2
                    ws_descarga_sal['O30'] = qtd2
                else :
                    ws_descarga_sal['D30'] = " "
                    ws_descarga_sal['H30'] = " "
                    ws_descarga_sal['K30'] = " "
                    ws_descarga_sal['O30'] = " "
                if checkbox_lote3_var.get() == 1:
                    ws_descarga_sal['D32'] = lote3
                    ws_descarga_sal['H32'] = nf3
                    ws_descarga_sal['K32'] = peso3
                    ws_descarga_sal['O32'] = qtd3
                else :
                    ws_descarga_sal['D32'] = " "
                    ws_descarga_sal['H32'] = " "
                    ws_descarga_sal['K32'] = " "
                    ws_descarga_sal['O32'] = " "
            

            # Salva a planilha
            wb.save('dados.xlsx')
            messagebox.showinfo("Sucesso", "Dados salvos com sucesso!")

        else:
            messagebox.showinfo("Erro", "Preencha todos os campos obrigatórios.")
    except ValueError as e:
        messagebox.showerror("Erro de valor", "Nos campos de QTD e Peso devese colocar exclusivamente numeros")
    except PermissionError as e:
        messagebox.showerror("Erro de permissão", f"Permissão negada: {e}. Verifique se o arquivo está aberto em outro programa.")
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")

def salvarEmb(date,horario,nome,telefone,placa,tipo,trans,forn,qtdtotalEmb,nfembalagem1,nfpaleteEmb1,codprod1,qtdpaleteEmb1,valEmb1,nomeprod1,contUnid1,lotef1,pesoEmb1,nfembalagem2,nfpaleteEmb2,codprod2,qtdpaleteEmb2,valEmb2,nomeprod2,contUnid2,lotef2,pesoEmb2
              ,nfembalagem3,nfpaleteEmb3,codprod3,qtdpaleteEmb3,valEmb3,nomeprod3,contUnid3,lotef3,pesoEmb3,nfembalagem4,nfpaleteEmb4,codprod4,qtdpaleteEmb4,valEmb4,nomeprod4,contUnid4,lotef4,pesoEmb4
              ,nfembalagem5,nfpaleteEmb5,codprod5,qtdpaleteEmb5,valEmb5,nomeprod5,contUnid5,lotef5,pesoEmb5,nfembalagem6,nfpaleteEmb6,codprod6,qtdpaleteEmb6,valEmb6,nomeprod6,contUnid6,lotef6,pesoEmb6):
    try:
        if date and horario and nome and telefone and placa and tipo and trans and forn:
            dados_EmbPlan = {'Movimento':['ENTRADA'],'EMISSÃO NF':[date],'PLACA':[placa],'TRANSPORTADOR':[trans],'MATERIAL':[nomeprod1],'TIPO DE PRODUTO':['EMBALAGEM'],'FORNECEDOR':[forn],'NF FORNCEDOR':[nfembalagem1],' QTDA ITENS ':[contUnid1],'QTD PALLET':[qtdpaleteEmb1],'NF PALLET':[nfpaleteEmb1],'LOTE 1 FORNECEDOR':[lotef1],'VALIDADE':[valEmb1],'PESO (KG)':[pesoEmb1]}
            df_dados_EmbPlan = pd.DataFrame(dados_EmbPlan)

            if nfembalagem2 != None:
                df_dados_EmbPlan = pd.concat([df_dados_EmbPlan,pd.DataFrame([{'Movimento':'ENTRADA','EMISSÃO NF':date,'PLACA':placa,'TRANSPORTADOR':trans,'MATERIAL':nomeprod2,'TIPO DE PRODUTO':'EMBALAGEM','FORNECEDOR':forn,'NF FORNCEDOR':nfembalagem2,' QTDA ITENS ':contUnid2,'QTD PALLET':qtdpaleteEmb2,'NF PALLET':nfpaleteEmb2,'LOTE 1 FORNECEDOR':lotef2,'VALIDADE':valEmb2,'PESO (KG)':pesoEmb2}])])

            if nfembalagem3 != None:
                df_dados_EmbPlan = pd.concat([df_dados_EmbPlan,pd.DataFrame([{'Movimento':'ENTRADA','EMISSÃO NF':date,'PLACA':placa,'TRANSPORTADOR':trans,'MATERIAL':nomeprod3,'TIPO DE PRODUTO':'EMBALAGEM','FORNECEDOR':forn,'NF FORNCEDOR':nfembalagem3,' QTDA ITENS ':contUnid3,'QTD PALLET':qtdpaleteEmb3,'NF PALLET':nfpaleteEmb3,'LOTE 1 FORNECEDOR':lotef3,'VALIDADE':valEmb3,'PESO (KG)':pesoEmb3}])])

            if nfembalagem4 != None:
                df_dados_EmbPlan = pd.concat([df_dados_EmbPlan,pd.DataFrame([{'Movimento':'ENTRADA','EMISSÃO NF':date,'PLACA':placa,'TRANSPORTADOR':trans,'MATERIAL':nomeprod4,'TIPO DE PRODUTO':'EMBALAGEM','FORNECEDOR':forn,'NF FORNCEDOR':nfembalagem4,' QTDA ITENS ':contUnid4,'QTD PALLET':qtdpaleteEmb4,'NF PALLET':nfpaleteEmb4,'LOTE 1 FORNECEDOR':lotef4,'VALIDADE':valEmb4,'PESO (KG)':pesoEmb4}])])
            
            if nfembalagem5 != None:
                df_dados_EmbPlan = pd.concat([df_dados_EmbPlan,pd.DataFrame([{'Movimento':'ENTRADA','EMISSÃO NF':date,'PLACA':placa,'TRANSPORTADOR':trans,'MATERIAL':nomeprod5,'TIPO DE PRODUTO':'EMBALAGEM','FORNECEDOR':forn,'NF FORNCEDOR':nfembalagem5,' QTDA ITENS ':contUnid5,'QTD PALLET':qtdpaleteEmb5,'NF PALLET':nfpaleteEmb5,'LOTE 1 FORNECEDOR':lotef5,'VALIDADE':valEmb5,'PESO (KG)':pesoEmb5}])])
            
            if nfembalagem6 != None:
                df_dados_EmbPlan = pd.concat([df_dados_EmbPlan,pd.DataFrame([{'Movimento':'ENTRADA','EMISSÃO NF':date,'PLACA':placa,'TRANSPORTADOR':trans,'MATERIAL':nomeprod6,'TIPO DE PRODUTO':'EMBALAGEM','FORNECEDOR':forn,'NF FORNCEDOR':nfembalagem6,' QTDA ITENS ':contUnid6,'QTD PALLET':qtdpaleteEmb6,'NF PALLET':nfpaleteEmb6,'LOTE 1 FORNECEDOR':lotef6,'VALIDADE':valEmb6,'PESO (KG)':pesoEmb6}])])
            
            if os.path.exists('dados.xlsx'):
                wb = load_workbook('dados.xlsx')
                if 'Embalagem Panilha' not in wb.sheetnames:
                    ws_dados_EmbPlan = wb.create_sheet("Embalagem Panilha")
                else:
                    ws_dados_EmbPlan = wb['Embalagem Panilha']
                
                for row in dataframe_to_rows(df_dados_EmbPlan, index=False, header=False):
                    ws_dados_EmbPlan.append(row)
                
                if 'Descarga Embalagem' not in wb.sheetnames:
                    ws_Descarga_Embalagem  = wb.create_sheet("Descarga Embalagem")
                else:
                    ws_Descarga_Embalagem = wb['Descarga Embalagem']
                    ws_Descarga_Embalagem['D8']=date
                    ws_Descarga_Embalagem['K8']=horario
                    ws_Descarga_Embalagem['R8']=placa
                    ws_Descarga_Embalagem['D10']=nome
                    ws_Descarga_Embalagem['R10']=telefone
                    ws_Descarga_Embalagem['D12']=tipo
                    ws_Descarga_Embalagem['O12']=trans
                    ws_Descarga_Embalagem['D14']=forn
                    ws_Descarga_Embalagem['R14']=qtdtotalEmb
                    ws_Descarga_Embalagem['D17']=nfembalagem1
                    ws_Descarga_Embalagem['D19']=codprod1
                    ws_Descarga_Embalagem['D21']=qtdpaleteEmb1
                    ws_Descarga_Embalagem['D23']=valEmb1
                    ws_Descarga_Embalagem['D25']=nomeprod1
                    ws_Descarga_Embalagem['D27']=contUnid1
                    ws_Descarga_Embalagem['D29']=lotef1
                    if nfembalagem2 != None:
                        ws_Descarga_Embalagem['E17']=nfembalagem2
                        ws_Descarga_Embalagem['E19']=codprod2
                        ws_Descarga_Embalagem['E21']=qtdpaleteEmb2
                        ws_Descarga_Embalagem['E23']=valEmb2
                        ws_Descarga_Embalagem['E25']=nomeprod2
                        ws_Descarga_Embalagem['E27']=contUnid2
                        ws_Descarga_Embalagem['E29']=lotef2
                    else:
                        ws_Descarga_Embalagem['E17']=" "
                        ws_Descarga_Embalagem['E19']=" "
                        ws_Descarga_Embalagem['E21']=" "
                        ws_Descarga_Embalagem['E23']=" "
                        ws_Descarga_Embalagem['E25']=" "
                        ws_Descarga_Embalagem['E27']=" "
                        ws_Descarga_Embalagem['E29']=" "
                    if nfembalagem3 != None:
                        ws_Descarga_Embalagem['J17']=nfembalagem3
                        ws_Descarga_Embalagem['J19']=codprod3
                        ws_Descarga_Embalagem['J21']=qtdpaleteEmb3
                        ws_Descarga_Embalagem['J23']=valEmb3
                        ws_Descarga_Embalagem['J25']=nomeprod3
                        ws_Descarga_Embalagem['J27']=contUnid3
                        ws_Descarga_Embalagem['J29']=lotef3
                    else:
                        ws_Descarga_Embalagem['J17']=" "
                        ws_Descarga_Embalagem['J19']=" "
                        ws_Descarga_Embalagem['J21']=" "
                        ws_Descarga_Embalagem['J23']=" "
                        ws_Descarga_Embalagem['J25']=" "
                        ws_Descarga_Embalagem['J27']=" "
                        ws_Descarga_Embalagem['J29']=" "
                    if nfembalagem4 != None:
                        ws_Descarga_Embalagem['M17']=nfembalagem4
                        ws_Descarga_Embalagem['M19']=codprod4
                        ws_Descarga_Embalagem['M21']=qtdpaleteEmb4
                        ws_Descarga_Embalagem['M23']=valEmb4
                        ws_Descarga_Embalagem['M25']=nomeprod4
                        ws_Descarga_Embalagem['M27']=contUnid4
                        ws_Descarga_Embalagem['M29']=lotef4
                    else:
                        ws_Descarga_Embalagem['M17']=" "
                        ws_Descarga_Embalagem['M19']=" "
                        ws_Descarga_Embalagem['M21']=" "
                        ws_Descarga_Embalagem['M23']=" "
                        ws_Descarga_Embalagem['M25']=" "
                        ws_Descarga_Embalagem['M27']=" "
                        ws_Descarga_Embalagem['M29']=" "
                    if nfembalagem5 != None:
                        ws_Descarga_Embalagem['P17']=nfembalagem5
                        ws_Descarga_Embalagem['P19']=codprod5
                        ws_Descarga_Embalagem['P21']=qtdpaleteEmb5
                        ws_Descarga_Embalagem['P23']=valEmb5
                        ws_Descarga_Embalagem['P25']=nomeprod5
                        ws_Descarga_Embalagem['P27']=contUnid5
                        ws_Descarga_Embalagem['P29']=lotef5
                    else:
                        ws_Descarga_Embalagem['P17']=" "
                        ws_Descarga_Embalagem['P19']=" "
                        ws_Descarga_Embalagem['P21']=" "
                        ws_Descarga_Embalagem['P23']=" "
                        ws_Descarga_Embalagem['P25']=" "
                        ws_Descarga_Embalagem['P27']=" "
                        ws_Descarga_Embalagem['P29']=" "
                    if nfembalagem6 != None:
                        ws_Descarga_Embalagem['R17']=nfembalagem6
                        ws_Descarga_Embalagem['R19']=codprod6
                        ws_Descarga_Embalagem['R21']=qtdpaleteEmb6
                        ws_Descarga_Embalagem['R23']=valEmb6
                        ws_Descarga_Embalagem['R25']=nomeprod6
                        ws_Descarga_Embalagem['R27']=contUnid6
                        ws_Descarga_Embalagem['R29']=lotef6
                    else:
                        ws_Descarga_Embalagem['R17']=" "
                        ws_Descarga_Embalagem['R19']=" "
                        ws_Descarga_Embalagem['R21']=" "
                        ws_Descarga_Embalagem['R23']=" "
                        ws_Descarga_Embalagem['R25']=" "
                        ws_Descarga_Embalagem['R27']=" "
                        ws_Descarga_Embalagem['R29']=" "
            else:
                messagebox.showerror("Erro","planilha não existe")
            wb.save('dados.xlsx')
            messagebox.showinfo("Sucesso", "Dados salvos com sucesso!")
        else:
            messagebox.showerror("Erro","preencha as informações")
    except FileNotFoundError as e:
        messagebox.showerror("Erro de Arquivo", f"Arquivo não encontrado: {e}")
    except PermissionError as e:
        messagebox.showerror("Erro de Permissão", f"Permissão negada: {e}. Verifique se o arquivo está aberto em outro programa.")
    except ValueError as e:
        messagebox.showerror("Erro de Valor", f"Dados inválidos: {e}")
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro inesperado: {e}")
